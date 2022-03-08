import pandas as pd
from tkinter import filedialog
from tkinter import Tk
import formulas as f
import os
from openpyxl import load_workbook, Workbook


class Setup:
    """
    Initial program setup

    Establish a root directory and save file name

    """

    def __init__(self):
        """
        Setup info for later export

        establishes the root directory and final output file name

        Args:
            none

        Returns:
            rootdir (string): directory for all major outputs
            savefile (string): name for final output (xlsx or csv, typically)

        Raises:
            none

        """

        filetype_options = ['csv', 'nd2', 'sample']
        self.filetype = f.Form.userinput('filetype', filetype_options)
        root = Tk()
        root.withdraw()
        self.rootdir = filedialog.askdirectory()
        print(self.rootdir)
        self.savefile = input('\nName your output file: ')
        self.parallel = f.Form.inputbool('\nProcess multiple files at once? (y/n): ')
        if self.parallel:
            self.brightmethod = 'auto'
            self.display = False
        else:
            brightmethod_options = ['manual', 'auto']
            self.brightmethod = f.Form.userinput('brigthness thresholding method',
                                                 brightmethod_options)
            self.display = f.Form.inputbool('\nDisplay figures? (y/n): ')

    def filelist(self):
        """
        List form of all files

        All filepaths that meet criteria for analysis

        Args:
            self

        Returns:
            self.filelist (list): all filepaths in one list

        Raises:
            none

        """

        self.filelist = []
        for subdir, dirs, files in os.walk(self.rootdir, topdown=False):
            for file in files:
                if file.endswith(self.filetype):
                    file = file.replace('\\', '/')
                    subdir = subdir.replace('\\', '/')
                    filepath = f'{subdir}/{file}'
                    self.filelist += [filepath]
        self.filelist.sort(reverse=False)

    def dfread(self, filepath):
        """
        Read in ND2 files

        loop through directory and find nd2 movie files

        Args:
            format (string): 3 character file format

        Returns:
            movie class object

        Raises:
            none

        """

        name = filepath.split('/')[-1][:-4]
        if self.filetype == 'csv':
            tracking = pd.read_csv(filepath, index_col=[0])
            tracking.rename(columns={'m2': 'Brightness'}, inplace=True)
            movie = Movie(name, filepath, tracks = tracking)
        if self.filetype == 'nd2':
            tracking = Track(filepath)
            movie = Movie(name, filepath, tracks = tracking.data)
        return movie



class Track:
    """
    Identify and link trajectories

    Use commercial, modified nearest neighbor algorithm to link particles

    """

    def __init__(self, im_path, quiet = False):
        """
        Track particles

        Use trackpy to identify and link individual particles

        Args:
            im_path (string): nd2 image file im_path
            quiet (bool): display text readout from trackpy

        Returns:
            trajectory data table as self.data

        Raises:
            Exception: restarts the loop (limit errors are oddly common???)

        """

        from pims import ND2Reader_SDK
        import trackpy as tp
        import h5py
        import os


        diameter = 5
        pix_mov = 10
        if quiet:
            tp.quiet()
        power_through_it = True
        attempts = 0
        while power_through_it:
            with ND2Reader_SDK(im_path) as movie:
                if os.path.exists('data.h5'):
                    os.remove('data.h5')
                low_mass = min([i.min() for i in movie])
                try:
                    with tp.PandasHDFStore('data.h5') as s:
                        print('\n^^^ This error is normal... unfortunately',
                              'Beginning Batch Processing', sep = 2*'\n')
                        tp.batch(movie, diameter, minmass = low_mass,
                                 processes = 'auto', output = s)
                        print('Batch Processing Complete',
                              'Beginning Trajectory Linking', sep = 2*'\n')
                        for linked in tp.link_df_iter(s, pix_mov):
                            s.put(linked)
                        trajs = pd.concat(iter(s))
                        power_through_it = False
                        print('Trajectory Linking Complete')
                except Exception:
                    print('Occassionally this part has issues, trying again')
                    attempts += 1
                    if attempts > 10:
                        power_through_it = False
        if os.path.exists('data.h5'):
            os.remove('data.h5')
        if trajs:
            trajs = trajs.rename(columns={'particle': 'Trajectory',
                                          'frame': 'Frame',
                                          'mass': 'Brightness'})
            trajs = f.Form.reorder(trajs, 'Trajectory', 0)
            trajs = f.Form.reorder(trajs, 'Frame', 1)
            trajs = f.Form.reorder(trajs, 'x', 2)
            trajs = f.Form.reorder(trajs, 'y', 3)
            trajs = f.Form.reorder(trajs, 'Brightness', 4)
            trajs.to_csv(f'{im_path[:-4]}_TP.csv')
            self.data = trajs
        else:
            print('utter failure')


class Movie:
    """
    Movie object class

    Extract and establish data from image stack (movie) file

    """

    def __init__(self, name, file,
                 tracks, frame_cutoff=5):
        """
        Establish movie file parameters

        Define metadata related to movie and camera

        Args:
            name (string): name of movies
            file (string): rootdir location
            frame_cutoff (int) : minimum frames for various analyses

        Returns:
            object metadata attributes

        Raises:
            none

        """

        self.name = {'Name': name}
        self.date = f.Find.identifiers(
            file, '/', 'Date Acquired', ['ymd'])
        self.protein_info = f.Find.identifiers(
            name, '_', 'Protein', ['grp', 'pdk'])
        self.gasket = f.Find.identifiers(file, '/', 'Gasket', ['gas'])
        self.replicate = {'Replicate': name[-2:]}
        self.ND = f.Find.identifiers(name, '_', 'ND Filter', ['nd'], '08')
        self.pixel_size = 0.000024
        self.framestep_size = 0.0217
        self.frame_cutoff = frame_cutoff
        self.fig_title = f.Form.catdict(self.protein_info, self.ND,
                                        self.gasket, self.replicate)
        self.df = tracks
        self.exportdata = self.name | self.date | self.gasket | \
                          self.replicate | self.protein_info | self.ND
        print(f'\nImage Name : {name}',
              self.date, self.gasket, self.replicate, self.ND,
              f'Initial Trajectory Count : {f.Calc.traj_count(self.df)}',
              sep='\n', end='\n'*2)

class Exports:
    """
    Export data

    Save and ammend exported data

    """

    def __init__(self, name, rootdir):
        """
        Export acquired data

        Build the dataframe and export all its data

        Args:
            name (string): typically the same as Setup.rootdir()
            export_df: takes keys (columns) from export data and creates empty
                dataframe
            rootdir (str): root directory

        Returns:
            none

        Raises:
            none

        """

        self.name = name
        self.exportpath = rootdir+f'\\{self.name}.xlsx'

    def start_df(self, exportdata):
        start = pd.DataFrame(columns=list(exportdata.keys()))
        self.export_df = start

    def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
        """
        Add row to Excel

        Use Excel Writer to add row to xlsx file

        Args:
            filename (str): Excel filename
            df (df): dataframe to append
            sheet_name (str): Excel file sheet tab name
            startrow (int): row to append data
            truncate_sheet (bool): shorten sheet
            **to_excel_kwargs (dict): additional arguments

        Returns:
            new excel file

        Raises:
            none

        """

        if not os.path.isfile(filename):
            df.to_excel(
                filename,
                sheet_name=sheet_name,
                startrow=startrow if startrow is not None else 0,
                **to_excel_kwargs)
            return
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')
        options = {}
        options['strings_to_formulas'] = False
        options['strings_to_urls'] = False
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a',
                                options=options)
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        if startrow is None:
            startrow = 0
        df.to_excel(writer, sheet_name, startrow=startrow,
                    **to_excel_kwargs)
        writer.save()

    def writer(path, df, mode='w', engine='openpyxl', if_sheet_exists=None,
               sheet_name='Raw'):
        with pd.ExcelWriter(path, mode=mode, engine=engine,
                            if_sheet_exists=if_sheet_exists) as writer:
            df.to_excel(writer, sheet_name=sheet_name)

    def build_df(self, dict):
        """
        Generates the export dataframe

        Appends dictionary values to export_df

        Args:
            dict (dictionary): dictionary of movie kinetics/values

        Returns:
            export_df (df): dataframe containing data from all movies

        Raises:
            none

        """

        new_row = pd.DataFrame(dict, index = [0])
        self.export_df = pd.concat([self.export_df, new_row])

    def csv(self, df, rootdir):
        """
        Export dataframe as csv file

        can cause errors where the main script tries to read this file as image

        Args:
            df (df): final dataframe with all calculated data
            rootdir (string): directory location of save file

        Returns:
            csv file

        Raises:
            none

        """

        full_file = script.rootdir+f'\\{self.name}'
        self.fidf = df
        csv_doc = df.to_csv(full_file+'.xlsx', index=False)
        return csv_doc

    def xlsx(self, script, df):
        """
        Export dataframe as xlsx file

        Tabs can be generated within this file to make it more digestable

        Args:
            script: script class object
            df (df): final dataframe with all calculated data

        Returns:
            xlsx file

        Raises:
            none

        """

        full_file = script.rootdir+f'\\{self.name}'
        self.fidf = df
        try:
            excel_doc = df.to_excel(full_file+'.xlsx', index=False, sheet_name='Raw')
            print('Export Successful')
            return excel_doc
        except Exception:
            print('Export Failed')

    def figure(rootdir, name):
        exdir = rootdir+'\Figures\\'
        full = exdir + ''
        pass
    # def xlsx_write(self, kw):
    #     sheet_df = self.fdf.filter(regex=kw)
    #     with pd.ExcelWriter(self.full_file+'.xlsx') as writer:
    #         sheet_df.to_excel(writer, sheet_name=kw, index=False)
